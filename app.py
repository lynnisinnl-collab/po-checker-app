import os
import json
import io
import time  
import re
from datetime import datetime
import pandas as pd
from google import genai  
from google.genai import types  
import openpyxl
from openpyxl.styles import PatternFill
import streamlit as st
import pypdf  # Requirement: pip install pypdf

# ==========================================
# APP CONFIGURATION & UI SETUP (ALL ENGLISH)
# ==========================================
st.set_page_config(page_title="PO Checker AI", layout="wide")
st.title("📦 Purchase Order Checking Assistant")
st.write("Upload your System Master Data and PO PDFs to automatically generate a flagged discrepancy report with matching confirmation notes (Supports Split Deliveries & Loose Text/Line Matching).")

# Securely fetch API key
if "GOOGLE_API_KEY" in st.secrets:
    GOOGLE_API_KEY = st.secrets["GOOGLE_API_KEY"]
else:
    st.error("🔑 Google API Key is not set! Please configure it within Streamlit Secrets.")
    st.stop()

client = genai.Client(api_key=GOOGLE_API_KEY)
OUTPUT_FILENAME = "PO_Checking_Report.xlsx"

# Advanced matching cleanup helper to handle revision suffixes (e.g., rev01, rev03)
def clean_key(val):
    if pd.isna(val) or val is None: return ""
    s = str(val).replace(" ", "").replace("-", "").replace(".", "").lower().strip()
    if 'rev' in s:
        s = s.split('rev')[0].strip()
    return s

# ULTIMATE SEMANTIC TEXT CLEANER: Translates word-numbers (four) to digit-numbers (4) and unifies '+' with '&'
def clean_description_semantic(val):
    if pd.isna(val) or val is None: return ""
    s = str(val).lower().strip()
    
    s = s.replace('+', '&').replace('and', '&').replace('en', '&')
    
    word_to_digit = {
        'zero': '0', 'one': '1', 'two': '2', 'three': '3', 'four': '4',
        'five': '5', 'six': '6', 'seven': '7', 'eight': '8', 'nine': '9', 'ten': '10'
    }
    for word, digit in word_to_digit.items():
        s = re.sub(r'\b' + word + r'\b', digit, s)
        
    s = re.sub(r'[^a-z0-9&]', '', s)
    return s

# Standardize line numbers (e.g., '0009' -> '9', '090/000' -> '9')
def clean_line_num(val):
    if pd.isna(val) or val is None: return ""
    s = str(val).strip().split('-')[0].split('/')[0].strip()
    return s.lstrip('0')

# Robust parsing tool to strip extra zeros (,0000) and isolate units (ST, pack, EA)
def parse_qty_and_unit(v):
    if pd.isna(v) or v is None:
        return None, None
    s = str(v).strip()
    
    match = re.match(r'^([\d\s.,]+)(.*)$', s)
    if match:
        num_part = match.group(1).strip()
        unit_part = match.group(2).strip()
        
        if ',' in num_part and '.' not in num_part:
            parts = num_part.split(',')
            if len(parts) == 2 and parts[1] == '0000':
                num_part = parts[0]
            else:
                num_part = num_part.replace(',', '.')
        elif '.' in num_part and ',' not in num_part:
            parts = num_part.split('.')
            if len(parts) == 2 and parts[1] == '0000':
                num_part = parts[0]
                
        num_part = num_part.replace(' ', '')
        
        try:
            val = float(num_part)
            if val.is_integer():
                val = int(val)
            return val, unit_part if unit_part else None
        except:
            return None, None
    return None, None

# Dynamic Price Normalizer handling 'à', 'per X', '€', and multlingual variations
def normalize_price(v):
    if pd.isna(v) or v is None: return None
    s = str(v).lower().strip()
    s = s.replace('à', '')
    
    factor = 1.0
    match_factor = re.search(r'(?:per|/)\s*(\d+)', s)
    if match_factor:
        factor = float(match_factor.group(1))
        
    s = re.sub(r'(?:per|/)\s*\d+\s*[a-z]*', '', s)
    
    for text_to_remove in ['st.', 'st', '€', 'eur', 'piece', 'pcs', 'ea', 'pack']:
        s = s.replace(text_to_remove, '')
    s = s.strip()
    
    if not s: return None
    if s.startswith(','): s = '0' + s
        
    s = s.replace(' ', '')
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'): s = s.replace('.', '').replace(',', '.')
        else: s = s.replace(',', '')
    elif ',' in s and '.' not in s:
        s = s.replace(',', '.')
        
    try:
        unit_val = float(s)
        return round(unit_val / factor, 4)
    except:
        return None

# Smart date parser supporting dot (.), slash (/), and hyphen (-) formats natively
def parse_date_to_custom_format(v):
    if pd.isna(v) or v is None: return ""
    s = str(v).replace(':', '').replace('Delivery Date', '').strip()
    s = s.replace('.', '/').replace('-', '/')  
    s = s.split(',')[0].strip().split()[0]
    
    for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y'):
        try: 
            return pd.to_datetime(s, format=fmt).strftime('%d/%m/%Y')
        except: 
            continue
    return str(v).strip()

# ==========================================
# UI FILE UPLOADERS
# ==========================================
excel_file = st.file_uploader("👉 Step 1: Upload System Master Data (Excel or CSV)", type=["xlsx", "xls", "csv"], key="master_data_excel_csv")
pdf_files = st.file_uploader("👉 Step 2: Upload PO PDF file(s)", type=["pdf"], accept_multiple_files=True, key="po_pdf_files_list")

# ==========================================
# APP PROCESSING LOGIC
# ==========================================
if excel_file and pdf_files:
    if st.button("🚀 Run AI Verification Report"):
        
        if excel_file.name.lower().endswith('.csv'):
            df_excel = pd.read_csv(excel_file, dtype=str)
        else:
            df_excel = pd.read_excel(excel_file, dtype=str)
            
        df_excel.columns = df_excel.columns.astype(str).str.strip()

        item_col_name = next((c for c in df_excel.columns if c.lower() in ['item', 'item number', 'material', 'part no', 'part number']), 'Item')
        if item_col_name not in df_excel.columns and len(df_excel.columns) > 0:
            item_col_name = df_excel.columns[0]

        unnamed_col = next((c for c in df_excel.columns if 'Unnamed' in c), 'Description_Extracted')
        if unnamed_col not in df_excel.columns:
            df_excel[unnamed_col] = None

        schedule_item_schema = types.Schema(
            type=types.Type.OBJECT,
            properties={
                "Split_Quantity": types.Schema(type=types.Type.STRING, description="The specific partial quantity for this split date (e.g. '25 st')"),
                "Required_Date": types.Schema(type=types.Type.STRING, description="The delivery date for this specific batch (e.g. '7-8-2026')")
            },
            required=["Split_Quantity", "Required_Date"]
        )

        po_item_schema = types.Schema(
            type=types.Type.OBJECT,
            properties={
                "Line": types.Schema(type=types.Type.STRING, description="Sequence position row index or item index (e.g. '010/000' or '090/000')"),
                "Item": types.Schema(type=types.Type.STRING, description="The primary item number listed under standard column (e.g. 511154)"),
                "Customer_Item": types.Schema(type=types.Type.STRING, description="The buyer's part number from notes labeled as 'Uw teknr:' (Leave blank if missing entirely)"),
                "Description": types.Schema(type=types.Type.STRING, description="Product item text description (e.g. 'Set Point 4 & 5 disk vlgs tek')"),
                "Unit_Price": types.Schema(type=types.Type.STRING, description="Price text including conditions (e.g., '€ 3.457,44 per 1 st')"),
                "Deliveries": types.Schema(
                    type=types.Type.ARRAY, 
                    items=schedule_item_schema, 
                    description="List of all scheduled delivery dates and split quantities for this single line item."
                )
            },
            required=["Item", "Unit_Price", "Deliveries"]
        )

        final_response_schema = types.Schema(
            type=types.Type.ARRAY,
            items=po_item_schema
        )

        prompt = """
        You are an elite Purchase Order parsing specialist handling complex multi-lingual layouts with split delivery dates and irregular descriptions.
        
        CRITICAL EXTRACTION RULES:
        1. EXTRACT ALL ITEMS NATIVELY:
           - Even if an item does NOT contain a customer part number ('Uw teknr'), you MUST extract it completely!
           - Capture the row 'Line' (e.g. '090/000'), 'Item' (e.g. '511154'), and 'Description' (e.g. 'Set Point 4 & 5 disk vlgs tek').
        2. SPLIT DELIVERIES / VERZENDSCHEMA:
           - Scan underneath the line item for single or multiple delivery blocks. If a delivery date exists without an explicit split quantity listed next to it, assume it takes the full or default quantity for that line.
        """

        all_po_items = []
        progress_bar = st.progress(0)
        quota_exhausted = False
        
        for idx, pdf_file in enumerate(pdf_files):
            st.write(f"🔍 AI is processing file: **{pdf_file.name}**")
            try:
                pdf_file.seek(0)
                pdf_bytes = pdf_file.read()
                if len(pdf_bytes) == 0: continue

                pdf_reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
                total_pages = len(pdf_reader.pages)
                pages_per_chunk = 2 
                
                for start_page in range(0, total_pages, pages_per_chunk):
                    end_page = min(start_page + pages_per_chunk, total_pages)
                    st.write(f"   📄 Parsing pages {start_page + 1} to {end_page} (Total: {total_pages} pages)...")
                    
                    pdf_writer = pypdf.PdfWriter()
                    for p_idx in range(start_page, end_page):
                        pdf_writer.add_page(pdf_reader.pages[p_idx])
                    
                    chunk_buffer = io.BytesIO()
                    pdf_writer.write(chunk_buffer)
                    chunk_bytes = chunk_buffer.getvalue()

                    max_retries = 4
                    retry_delay = 4  
                    response = None
                    
                    for attempt in range(max_retries):
                        try:
                            response = client.models.generate_content(
                                model='gemini-2.5-flash',
                                contents=[
                                    types.Part.from_bytes(data=chunk_bytes, mime_type='application/pdf'),
                                    prompt
                                ],
                                config=types.GenerateContentConfig(
                                    response_mime_type="application/json",
                                    response_schema=final_response_schema,
                                    temperature=0.0
                                )
                            )
                            break  
                        except Exception as api_err:
                            err_msg = str(api_err)
                            if "429" in err_msg or "RESOURCE_EXHAUSTED" in err_msg:
                                st.error("🛑 **Daily Quota Limit Reached (429)!**")
                                quota_exhausted = True
                                break
                            if "503" in err_msg or "UNAVAILABLE" in err_msg:
                                if attempt < max_retries - 1:
                                    time.sleep(retry_delay)
                                    retry_delay *= 2  
                                    continue
                            raise api_err  
                    
                    if quota_exhausted: st.stop()
                        
                    if response and response.text:
                        try:
                            items_data = json.loads(response.text.strip())
                            for item in items_data:
                                item['PO_Source_File'] = pdf_file.name
                                all_po_items.append(item)
                        except json.JSONDecodeError:
                            st.warning(f"⚠️ Failed to parse JSON for pages {start_page+1}-{end_page}, skipping chunk.")
                    time.sleep(2)
                        
            except Exception as e:
                st.error(f"❌ Failed to parse {pdf_file.name}: {e}")
            
            progress_bar.progress((idx + 1) / len(pdf_files))

        if not all_po_items:
            st.error("❌ No data was extracted from your PDF items. Processing stopped.")
            st.stop()

        df_po = pd.DataFrame(all_po_items)
        st.write("🔄 Aligning rows via translation-aware semantic routing rules...")

        structured_rows = []
        excel_row_blocks_meta = [] 

        for idx, row in df_excel.iterrows():
            excel_item = str(row.get(item_col_name, '')).strip()
            excel_key = clean_key(excel_item)
            ex_line = clean_line_num(row.get('Line'))
            
            excel_desc_cleaned = clean_description_semantic(row.get(unnamed_col, ''))
            
            excel_side_row = {col: row[col] for col in df_excel.columns}
            excel_side_row['Data Block Source'] = 'GloviaG2'
            
            if 'Order Quantity' in excel_side_row:
                ex_qty_num, _ = parse_qty_and_unit(excel_side_row['Order Quantity'])
                if ex_qty_num is not None: excel_side_row['Order Quantity'] = str(ex_qty_num)
            
            if 'Unit Price' in excel_side_row:
                ex_price_num = normalize_price(excel_side_row['Unit Price'])
                if ex_price_num is not None: excel_side_row['Unit Price'] = str(ex_price_num)
            
            if 'Required Date/Time' in excel_side_row:
                excel_side_row['Required Date/Time'] = parse_date_to_custom_format(excel_side_row['Required Date/Time'])
            
            matched_po_item = None
            if not df_po.empty:
                candidates = []
                for po_idx, po_row in df_po.iterrows():
                    po_key_p = clean_key(po_row.get('Item', ''))
                    po_key_c = clean_key(po_row.get('Customer_Item', ''))
                    po_desc_cleaned = clean_description_semantic(po_row.get('Description', ''))
                    po_line_cleaned = clean_line_num(po_row.get('Line', ''))
                    
                    is_num_match = (excel_key and ((po_key_p and (po_key_p in excel_key or excel_key in po_key_p)) or 
                                                   (po_key_c and (po_key_c in excel_key or excel_key in po_key_c))))
                    
                    is_line_index_match = (ex_line and po_line_cleaned and ex_line == po_line_cleaned)
                    is_semantic_desc_match = (excel_desc_cleaned and po_desc_cleaned and \
                                              (excel_desc_cleaned in po_desc_cleaned or po_desc_cleaned in excel_desc_cleaned or \
                                               excel_desc_cleaned[:10] in po_desc_cleaned or po_desc_cleaned[:10] in excel_desc_cleaned))
                    
                    if is_num_match or (is_line_index_match and is_semantic_desc_match) or (is_line_index_match and not excel_key):
                        candidates.append(po_row)
                
                if candidates:
                    matched_po_item = candidates[0]
                    for cand in candidates:
                        if clean_line_num(cand.get('Line')) == ex_line:
                            matched_po_item = cand
                            break

            block_start_index = len(structured_rows)
            
            if matched_po_item is not None:
                deliveries = matched_po_item.get('Deliveries', [])
                if not isinstance(deliveries, list) or len(deliveries) == 0:
                    deliveries = [{"Split_Quantity": matched_po_item.get('Order_Quantity', '0'), "Required_Date": matched_po_item.get('Required_Date', '')}]
                
                total_pdf_item_qty = 0
                for d in deliveries:
                    q_num, _ = parse_qty_and_unit(d.get('Split_Quantity', '0'))
                    if q_num: total_pdf_item_qty += q_num
                if total_pdf_item_qty == 0:
                    total_pdf_item_qty, _ = parse_qty_and_unit(excel_side_row.get('Order Quantity', '0'))

                structured_rows.append(excel_side_row)

                for d_idx, deliv in enumerate(deliveries):
                    pdf_side_row = {col: None for col in df_excel.columns}
                    pdf_side_row['Data Block Source'] = 'PDF'
                    
                    # Show actual extracted Part Number from PDF instead of blindly copying Excel
                    pdf_item = str(matched_po_item.get('Item', '')).strip()
                    pdf_cust = str(matched_po_item.get('Customer_Item', '')).strip()
                    if pdf_cust and pdf_cust.lower() not in ['none', 'null', '']:
                        pdf_side_row[item_col_name] = f"{pdf_item} / {pdf_cust}"
                    else:
                        pdf_side_row[item_col_name] = pdf_item
                    
                    if 'Line' in df_excel.columns: pdf_side_row['Line'] = matched_po_item.get('Line')
                    
                    raw_price = matched_po_item.get('Unit_Price')
                    calc_price = normalize_price(raw_price)
                    if 'Unit Price' in df_excel.columns:
                        pdf_side_row['Unit Price'] = str(calc_price) if calc_price is not None else raw_price
                        
                    pdf_side_row[unnamed_col] = matched_po_item.get('Description', 'No Description')
                    if 'Notes' in df_excel.columns: 
                        pdf_side_row['Notes'] = f"Extracted from PDF: {matched_po_item.get('PO_Source_File', '')} [Batch {d_idx+1}]"
                    
                    raw_split_qty = deliv.get('Split_Quantity')
                    clean_split_qty, extracted_unit = parse_qty_and_unit(raw_split_qty)
                    if (clean_split_qty is None or clean_split_qty == 0) and len(deliveries) == 1:
                        clean_split_qty = total_pdf_item_qty
                        
                    if 'Order Quantity' in df_excel.columns:
                        pdf_side_row['Order Quantity'] = str(clean_split_qty) if clean_split_qty is not None else raw_split_qty
                        
                    for um_col in ['UM', 'Stock UM', 'In Stock UM']:
                        if um_col in df_excel.columns:
                            pdf_side_row[um_col] = extracted_unit if extracted_unit else excel_side_row.get(um_col)

                    formatted_deliv_date = parse_date_to_custom_format(deliv.get('Required_Date'))
                    if 'Required Date/Time' in df_excel.columns:
                        pdf_side_row['Required Date/Time'] = formatted_deliv_date

                    conf_text = ""
                    if formatted_deliv_date:
                        today_str = datetime.now().strftime("%d%m")
                        ddmmyy = ""
                        if '/' in formatted_deliv_date:
                            parts = formatted_deliv_date.split('/')
                            if len(parts) >= 3: ddmmyy = parts[0] + parts[1] + parts[2][-2:]
                        conf_text = f"{today_str} lla dd conf. {ddmmyy}"
                    
                    pdf_side_row['Confirmation Note'] = conf_text
                    structured_rows.append(pdf_side_row)
                
                block_end_index = len(structured_rows)
                excel_row_blocks_meta.append({
                    "type": "MATCHED",
                    "start": block_start_index,
                    "end": block_end_index,
                    "total_pdf_qty": total_pdf_item_qty
                })
            else:
                if 'Notes' in df_excel.columns: excel_side_row['Notes'] = ""
                excel_side_row['Confirmation Note'] = ""
                structured_rows.append(excel_side_row)
                
                pdf_side_row = {col: None for col in df_excel.columns}
                pdf_side_row['Data Block Source'] = 'PDF'
                pdf_side_row[item_col_name] = excel_item
                if 'Notes' in df_excel.columns: pdf_side_row['Notes'] = "Not found in PO PDF"
                for um_col in ['UM', 'Stock UM', 'In Stock UM']:
                    if um_col in df_excel.columns: pdf_side_row[um_col] = excel_side_row.get(um_col)
                structured_rows.append(pdf_side_row)
                
                block_end_index = len(structured_rows)
                excel_row_blocks_meta.append({
                    "type": "MISSING",
                    "start": block_start_index,
                    "end": block_end_index,
                    "total_pdf_qty": 0
                })

            blank_row = {col: None for col in df_excel.columns}
            blank_row['Data Block Source'] = None
            blank_row['Confirmation Note'] = None
            structured_rows.append(blank_row)

        df_final = pd.DataFrame(structured_rows)
        core_cols = [c for c in df_final.columns if c not in ['Data Block Source', 'Confirmation Note']]
        cols = ['Data Block Source'] + core_cols + ['Confirmation Note']
        df_final = df_final[cols]
        df_final.to_excel(OUTPUT_FILENAME, index=False)

        # ==========================================
        # Style and Highlight Output
        # ==========================================
        wb = openpyxl.load_workbook(OUTPUT_FILENAME)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        idx_item = headers.index(item_col_name) + 1 if item_col_name in headers else None
        idx_desc = headers.index(unnamed_col) + 1 if unnamed_col in headers else None
        idx_qty = headers.index('Order Quantity') + 1 if 'Order Quantity' in headers else None
        idx_price = headers.index('Unit Price') + 1 if 'Unit Price' in headers else None
        idx_date = headers.index('Required Date/Time') + 1 if 'Required Date/Time' in headers else None

        fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        fill_light_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        fill_yellow = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        
        for block in excel_row_blocks_meta:
            start_r = block["start"] + 2 
            end_r = block["end"] + 1
            glovia_row_num = start_r
            
            if block["type"] == "MISSING":
                for r in range(start_r, end_r):
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=r, column=c).fill = fill_yellow
                continue
            
            for r in range(start_r, end_r):
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = fill_light_gray
            
            # Fetch Baseline Glovia Master Values
            ex_item_val = clean_key(ws.cell(row=glovia_row_num, column=idx_item).value if idx_item else "")
            ex_desc_val = clean_description_semantic(ws.cell(row=glovia_row_num, column=idx_desc).value if idx_desc else "")
            ex_qty_val, _ = parse_qty_and_unit(ws.cell(row=glovia_row_num, column=idx_qty).value if idx_qty else 0)
            ex_price_val = normalize_price(ws.cell(row=glovia_row_num, column=idx_price).value if idx_price else 0)
            
            ex_date_val = ""
            if idx_date:
                raw_d = ws.cell(row=glovia_row_num, column=idx_date).value
                if raw_d is not None and str(raw_d).strip().lower() not in ["none", "nan", ""]:
                    ex_date_val = str(raw_d).strip()

            qty_discrepancy = (ex_qty_val != block["total_pdf_qty"])

            # Check Discrepancies and Highlight BOTH rows red if a mismatch occurs
            for pdf_row_num in range(glovia_row_num + 1, end_r):
                
                # 1. ITEM NUMBER HIGHLIGHT
                if idx_item:
                    i_cell = ws.cell(row=pdf_row_num, column=idx_item)
                    pdf_item_val = clean_key(i_cell.value)
                    # If Excel code is completely missing from the captured PDF code block
                    if ex_item_val and pdf_item_val and ex_item_val not in pdf_item_val:
                        ws.cell(row=glovia_row_num, column=idx_item).fill = fill_red
                        i_cell.fill = fill_red

                # 2. DESCRIPTION (ITEM NAME) HIGHLIGHT
                if idx_desc:
                    desc_cell = ws.cell(row=pdf_row_num, column=idx_desc)
                    pdf_desc_val = clean_description_semantic(desc_cell.value)
                    
                    is_desc_match = False
                    if not ex_desc_val and not pdf_desc_val:
                        is_desc_match = True
                    elif ex_desc_val and pdf_desc_val:
                        if ex_desc_val in pdf_desc_val or pdf_desc_val in ex_desc_val:
                            is_desc_match = True
                    
                    if not is_desc_match and desc_cell.value is not None:
                        ws.cell(row=glovia_row_num, column=idx_desc).fill = fill_red
                        desc_cell.fill = fill_red

                # 3. QUANTITY HIGHLIGHT
                if idx_qty and qty_discrepancy:
                    ws.cell(row=glovia_row_num, column=idx_qty).fill = fill_red
                    ws.cell(row=pdf_row_num, column=idx_qty).fill = fill_red
                    
                # 4. PRICE HIGHLIGHT
                if idx_price:
                    p_cell = ws.cell(row=pdf_row_num, column=idx_price)
                    p_val = normalize_price(p_cell.value)
                    if p_val is not None and ex_price_val != p_val:
                        ws.cell(row=glovia_row_num, column=idx_price).fill = fill_red
                        p_cell.fill = fill_red
                        
                # 5. DATE HIGHLIGHT 
                if idx_date:
                    d_cell = ws.cell(row=pdf_row_num, column=idx_date)
                    pdf_date_val = ""
                    if d_cell.value is not None and str(d_cell.value).strip().lower() not in ["none", "nan", ""]:
                        pdf_date_val = str(d_cell.value).strip()
                        
                    if pdf_date_val != "" and ex_date_val != pdf_date_val:
                        ws.cell(row=glovia_row_num, column=idx_date).fill = fill_red
                        d_cell.fill = fill_red

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        st.success("🎉 Process Complete!")
        st.download_button(
            label="📥 Download Discrepancy Report",
            data=excel_buffer,
            file_name=OUTPUT_FILENAME,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("💡 Please upload both the Master Excel file and PO PDFs to begin.")
